import os
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NOME_ARQUIVO = "banco_de_senhas_automatizado.xlsx"


def aplicar_estilos_cabecalho(ws):
    headers = [
        "ID",
        "Nome / Responsável",
        "Senha (4 Dígitos)",
        "Data de Criação",
        "Prazo (Dias)",
        "Data Vencimento",
        "Status",
    ]

    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    larguras = {"A": 8, "B": 28, "C": 18, "D": 22, "E": 15, "F": 22, "G": 16}
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width


def aplicar_estilo_linha(ws, linha):
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    for c in range(1, 8):
        cell = ws.cell(row=linha, column=c)
        cell.alignment = left_align if c == 2 else center_align
        cell.border = thin_border


def carregar_e_normalizar_dados():
    if not os.path.exists(NOME_ARQUIVO):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Banco de Senhas"
        aplicar_estilos_cabecalho(ws)
        wb.save(NOME_ARQUIVO)
        return []

    try:
        wb = openpyxl.load_workbook(NOME_ARQUIVO)
    except Exception:
        return []

    ws = wb["Banco de Senhas"]
    registros = []
    hoje = datetime.now()

    for r in range(2, ws.max_row + 1):
        nome = ws.cell(row=r, column=2).value
        senha = ws.cell(row=r, column=3).value
        data_criacao_str = ws.cell(row=r, column=4).value
        prazo_dias = ws.cell(row=r, column=5).value

        if not nome or not senha:
            continue

        nome = str(nome).strip()
        senha = str(senha).replace("'", "").strip().zfill(4)

        try:
            prazo_dias = int(prazo_dias)
        except Exception:
            prazo_dias = 30

        try:
            if isinstance(data_criacao_str, datetime):
                data_criacao = data_criacao_str
            else:
                data_criacao = datetime.strptime(
                    str(data_criacao_str), "%Y-%m-%d %H:%M:%S"
                )
        except Exception:
            data_criacao = hoje

        data_vencimento = data_criacao + timedelta(days=prazo_dias)
        status = (
            "Ativa" if prazo_dias > 0 and data_vencimento >= hoje else "Expirada"
        )

        registros.append({
            "nome": nome,
            "senha": senha,
            "data_criacao": data_criacao,
            "prazo_dias": prazo_dias,
            "data_vencimento": data_vencimento,
            "status": status,
        })

    registros.sort(key=lambda x: x["data_criacao"])

    for i, reg in enumerate(registros, start=1):
        reg["id"] = i

    wb.close()
    return registros


def salvar_dados_limpos(registros):
    for i, reg in enumerate(registros, start=1):
        reg["id"] = i

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Banco de Senhas"

    aplicar_estilos_cabecalho(ws)

    for idx, reg in enumerate(registros, start=2):
        ws.cell(row=idx, column=1, value=reg["id"])
        ws.cell(row=idx, column=2, value=reg["nome"])

        c_senha = ws.cell(row=idx, column=3, value=reg["senha"])
        c_senha.number_format = "@"

        ws.cell(
            row=idx,
            column=4,
            value=reg["data_criacao"].strftime("%Y-%m-%d %H:%M:%S"),
        )
        ws.cell(row=idx, column=5, value=reg["prazo_dias"])
        ws.cell(
            row=idx,
            column=6,
            value=reg["data_vencimento"].strftime("%Y-%m-%d %H:%M:%S"),
        )
        ws.cell(row=idx, column=7, value=reg["status"])

        aplicar_estilo_linha(ws, idx)

    wb.save(NOME_ARQUIVO)


def listar_cadastros():
    registros = carregar_e_normalizar_dados()
    salvar_dados_limpos(registros)

    if not registros:
        print("\n📭 Nenhum cadastro encontrado.")
        return False

    print("\n" + "=" * 70)
    print(
        f"{'ID':<5} | {'NOME':<25} | {'SENHA':<8} | {'STATUS':<10} | VENCIMENTO"
    )
    print("=" * 70)

    for reg in registros:
        venc_str = reg["data_vencimento"].strftime("%Y-%m-%d %H:%M:%S")
        print(
            f"{reg['id']:<5} | {reg['nome']:<25} | {reg['senha']:<8} |"
            f" {reg['status']:<10} | {venc_str}"
        )

    print("=" * 70 + "\n")
    return True


def criar_nova_senha():
    try:
        registros = carregar_e_normalizar_dados()

        print("\n--- 🔑 CRIAR NOVA SENHA ---")
        nome = input("👤 Digite o NOME do responsável: ").strip()
        while not nome:
            print("⚠️ O nome não pode ficar em branco.")
            nome = input("👤 Digite o NOME do responsável: ").strip()

        prazo_input = input(
            "⏳ Digite o prazo em DIAS (ENTER para 30 dias): "
        ).strip()
        try:
            prazo_dias = int(prazo_input) if prazo_input else 30
        except ValueError:
            print("⚠️ Prazo inválido. Usando padrão de 30 dias.")
            prazo_dias = 30

        senhas_ativas = {r["senha"] for r in registros if r["status"] == "Ativa"}
        if len(senhas_ativas) >= 10000:
            print("❌ Todas as combinações possíveis estão em uso!")
            return

        while True:
            nova_senha_num = f"{random.randint(0, 9999):04d}"
            if nova_senha_num not in senhas_ativas:
                break

        agora = datetime.now()
        data_vencimento = agora + timedelta(days=prazo_dias)
        status_inicial = "Ativa" if prazo_dias > 0 else "Expirada"

        novo_registro = {
            "nome": nome,
            "senha": nova_senha_num,
            "data_criacao": agora,
            "prazo_dias": prazo_dias,
            "data_vencimento": data_vencimento,
            "status": status_inicial,
        }

        registros.append(novo_registro)
        salvar_dados_limpos(registros)

        registros_atualizados = carregar_e_normalizar_dados()
        id_final = next(
            r["id"] for r in registros_atualizados if r["senha"] == nova_senha_num
        )

        print(
            f"\n✅ Sucesso! ID [{id_final}] | Senha [{nova_senha_num}] gerada"
            f" para '{nome}'."
        )

    except PermissionError:
        print(
            "\n❌ ERRO: O arquivo Excel está aberto! Feche a planilha e tente"
            " novamente."
        )


def editar_informacoes():
    try:
        registros = carregar_e_normalizar_dados()
        if not registros:
            print("\n📭 Nenhum cadastro encontrado para editar.")
            return

        print("\n--- ✏️ EDITAR INFORMAÇÕES ---")
        listar_cadastros()

        id_busca = input("Digite o ID da pessoa que deseja editar: ").strip()
        if not id_busca.isdigit():
            print("⚠️ ID inválido.")
            return

        id_alvo = int(id_busca)
        registro = next((r for r in registros if r["id"] == id_alvo), None)

        if not registro:
            print(f"❌ Cadastro com ID {id_alvo} não foi encontrado.")
            return

        print(f"\nEditando registro de: {registro['nome']}")
        novo_nome = input(
            f"Novo nome (Aperte ENTER para manter '{registro['nome']}'): "
        ).strip()
        novo_prazo = input(
            "Adicionar/Alterar prazo em DIAS (Aperte ENTER para manter"
            f" {registro['prazo_dias']} dias): "
        ).strip()

        if novo_nome:
            registro["nome"] = novo_nome

        if novo_prazo.isdigit() or (
            novo_prazo.startswith("-") and novo_prazo[1:].isdigit()
        ):
            dias = int(novo_prazo)
            registro["prazo_dias"] = dias
            registro["data_vencimento"] = registro["data_criacao"] + timedelta(
                days=dias
            )
            registro["status"] = (
                "Ativa"
                if dias > 0 and registro["data_vencimento"] >= datetime.now()
                else "Expirada"
            )

        salvar_dados_limpos(registros)
        print(f"✅ Informações do ID {id_alvo} atualizadas com sucesso!")

    except PermissionError:
        print(
            "\n❌ ERRO: O arquivo Excel está aberto! Feche a planilha e tente"
            " novamente."
        )


def deletar_pessoa():
    try:
        registros = carregar_e_normalizar_dados()
        if not registros:
            print("\n📭 Nenhum cadastro encontrado para deletar.")
            return

        print("\n--- 🗑️ DELETAR REGISTRO ---")
        listar_cadastros()

        id_busca = input("Digite o ID da pessoa que deseja DELETAR: ").strip()
        if not id_busca.isdigit():
            print("⚠️ ID inválido.")
            return

        id_alvo = int(id_busca)
        registro = next((r for r in registros if r["id"] == id_alvo), None)

        if not registro:
            print(f"❌ Cadastro com ID {id_alvo} não foi encontrado.")
            return

        confirmar = (
            input(
                f"⚠️ Tem certeza que deseja deletar '{registro['nome']}' (ID"
                f" {id_alvo})? (S/N): "
            )
            .strip()
            .lower()
        )

        if confirmar in ["s", "sim"]:
            novos_registros = [r for r in registros if r["id"] != id_alvo]
            salvar_dados_limpos(novos_registros)
            print(
                f"✅ Registro de '{registro['nome']}' deletado e planilha"
                " reorganizada!"
            )
        else:
            print("❌ Operação cancelada.")

    except PermissionError:
        print(
            "\n❌ ERRO: O arquivo Excel está aberto! Feche a planilha e tente"
            " novamente."
        )


def menu_principal():
    while True:
        print("\n" + "=" * 45)
        print("      📋 PAINEL DE CONTROLE DE SENHAS")
        print("=" * 45)
        print("1. 🔑 Gerar Nova Senha")
        print("2. ✏️ Editar Nome ou Prazo de Alguém")
        print("3. 🗑️ Deletar Cadastro")
        print("4. 📊 Listar Todos os Cadastros")
        print("5. 🚪 Sair")
        print("=" * 45)

        opcao = input("Escolha uma opção (1-5): ").strip()

        if opcao == "1":
            criar_nova_senha()
        elif opcao == "2":
            editar_informacoes()
        elif opcao == "3":
            deletar_pessoa()
        elif opcao == "4":
            listar_cadastros()
        elif opcao == "5":
            print("\n👋 Sistema encerrado com sucesso!")
            break
        else:
            print("\n⚠️ Opção inválida.")


if __name__ == "__main__":
    menu_principal()