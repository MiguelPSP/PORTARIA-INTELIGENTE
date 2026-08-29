import base64
import os
import random
import urllib.parse
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
import streamlit as st

# Configuração da página
st.set_page_config(
    page_title="Interface Digital - Controle de Senhas",
    page_icon="🛡️",
    layout="wide",
)

# ---------------------------------------------------------
# ESTILIZAÇÃO CSS
# ---------------------------------------------------------
st.markdown(
    """
    <style>
        .stApp {
            background-color: #FFFFFF;
            color: #181818;
        }
        [data-testid="stSidebar"] {
            background-color: #181818 !important;
            border-right: 3px solid #DDBB01;
        }
        [data-testid="stSidebar"] *, 
        [data-testid="stSidebar"] p, 
        [data-testid="stSidebar"] span, 
        [data-testid="stSidebar"] label {
            color: #FFFFFF !important;
        }
        div.stButton > button:first-child, [data-testid="stDownloadButton"] > button {
            background-color: #181818 !important;
            color: #DDBB01 !important;
            font-weight: bold !important;
            border-radius: 6px !important;
            border: 2px solid #DDBB01 !important;
            padding: 10px 24px !important;
            transition: all 0.3s ease !important;
            width: 100% !important;
        }
        div.stButton > button:first-child:hover, [data-testid="stDownloadButton"] > button:hover {
            background-color: #DDBB01 !important;
            color: #000000 !important;
            box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.2);
        }
        .company-link {
            display: inline-block;
            color: #DDBB01 !important;
            text-decoration: none;
            font-weight: bold;
            font-size: 14px;
            padding: 10px 12px;
            border: 1px solid #DDBB01;
            border-radius: 6px;
            text-align: center;
            width: 100%;
            background-color: #181818;
            transition: 0.3s;
            margin-bottom: 8px;
        }
        .company-link:hover {
            background-color: #DDBB01;
            color: #000000 !important;
        }
        .whatsapp-btn {
            display: inline-block;
            color: #FFFFFF !important;
            background-color: #25D366;
            text-decoration: none;
            font-weight: bold;
            font-size: 14px;
            padding: 10px 12px;
            border-radius: 6px;
            text-align: center;
            width: 100%;
            transition: 0.3s;
            box-shadow: 0px 4px 6px rgba(0,0,0,0.3);
        }
        .whatsapp-btn:hover {
            background-color: #1EBE5D;
            color: #FFFFFF !important;
        }
        .highlight-yellow {
            color: #DDBB01;
            font-weight: bold;
        }
    </style>
""",
    unsafe_allow_html=True,
)

NOME_ARQUIVO = "banco_de_senhas_automatizado.xlsx"

# Configuração do WhatsApp
NUMERO_WHATSAPP = "5581986296128"
MENSAGEM_PADRAO = urllib.parse.quote(
    "Olá! Preciso de suporte no sistema da portaria Interface Digital."
)
LINK_WHATSAPP = f"https://wa.me/{NUMERO_WHATSAPP}?text={MENSAGEM_PADRAO}"


def salvar_dados(registros):
    """Força a gravação física imediata da lista no arquivo Excel."""
    for i, reg in enumerate(registros, start=1):
        reg["id"] = i

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Banco de Senhas"
    ws.append([
        "ID",
        "Nome / Responsável",
        "Senha (4 Dígitos)",
        "Data de Criação",
        "Prazo (Dias)",
        "Data Vencimento",
        "Status",
    ])

    for reg in registros:
        ws.append([
            reg["id"],
            reg["nome"],
            f"'{reg['senha']}",
            reg["data_criacao"].strftime("%Y-%m-%d %H:%M:%S"),
            reg["prazo_dias"],
            reg["data_vencimento"].strftime("%Y-%m-%d %H:%M:%S"),
            reg["status"],
        ])

    wb.save(NOME_ARQUIVO)


def inicializar_e_carregar_dados():
    """Garante a integridade do arquivo Excel e carrega os dados em memória."""
    if not os.path.exists(NOME_ARQUIVO):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Banco de Senhas"
        ws.append([
            "ID",
            "Nome / Responsável",
            "Senha (4 Dígitos)",
            "Data de Criação",
            "Prazo (Dias)",
            "Data Vencimento",
            "Status",
        ])
        wb.save(NOME_ARQUIVO)

    wb = openpyxl.load_workbook(NOME_ARQUIVO)
    ws = wb["Banco de Senhas"]
    registros = []
    hoje = datetime.now()

    for r in range(2, ws.max_row + 1):
        nome = ws.cell(row=r, column=2).value
        senha = ws.cell(row=r, column=3).value
        data_criacao_str = ws.cell(row=r, column=4).value
        prazo_dias = ws.cell(row=r, column=5).value

        if not nome or senha is None:
            continue

        nome = str(nome).strip()
        senha = str(senha).replace("'", "").strip().zfill(4)

        try:
            prazo_dias = int(prazo_dias)
            if prazo_dias > 30:
                prazo_dias = 30
        except Exception:
            prazo_dias = 30

        try:
            if isinstance(data_criacao_str, datetime):
                data_criacao = data_criacao_str
            else:
                data_criacao = datetime.strptime(
                    str(data_criacao_str), "%Y-%m-%d %H:%M:%S"
                )
        except Exception:
            data_criacao = hoje

        data_vencimento = data_criacao + timedelta(days=prazo_dias)
        status = (
            "Ativa" if prazo_dias > 0 and data_vencimento >= hoje else "Expirada"
        )

        registros.append({
            "nome": nome,
            "senha": senha,
            "data_criacao": data_criacao,
            "prazo_dias": prazo_dias,
            "data_vencimento": data_vencimento,
            "status": status,
        })

    registros.sort(key=lambda x: x["data_criacao"])
    wb.close()

    salvar_dados(registros)
    return registros


# Carrega e atualiza registros no início
registros = inicializar_e_carregar_dados()

# ---------------------------------------------------------
# BARRA LATERAL (BRANDING, DOWNLOAD & SUPORTE)
# ---------------------------------------------------------
with st.sidebar:
    if os.path.exists("logo.png"):
        with open("logo.png", "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode()

        st.markdown(
            f"""
            <div style="
                background-color: #FFFFFF; 
                padding: 12px; 
                border-radius: 8px; 
                text-align: center; 
                margin-bottom: 20px;
                box-shadow: 0px 4px 8px rgba(0,0,0,0.3);
            ">
                <img src="data:image/png;base64,{encoded_string}" style="max-width: 100%; height: auto; display: block; margin: 0 auto;">
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            "## <span class='highlight-yellow'>INTERFACE</span> DIGITAL",
            unsafe_allow_html=True,
        )

    st.markdown("---")

    menu = st.radio(
        "Navegação da Portaria:",
        [
            "📊 Painel de Controle",
            "🔑 Novo Cadastro de Senha",
            "✏️ Editar Cadastro",
            "🗑️ Cancelar / Remover Senha",
        ],
    )

    st.markdown("---")

    if os.path.exists(NOME_ARQUIVO):
        with open(NOME_ARQUIVO, "rb") as file:
            st.download_button(
                label="📥 Baixar Dados (.xlsx)",
                data=file,
                file_name=f"relatorio_senhas_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    st.markdown("---")

    st.markdown(
        f"""
        <a href="{LINK_WHATSAPP}" target="_blank" class="whatsapp-btn">
            💬 Suporte WhatsApp
        </a>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <a href="https://www.interfacedigital.com.br" target="_blank" class="company-link" style="margin-top: 8px;">
            🌐 Site Oficial
        </a>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        "<p style='text-align: center; color: #888888; font-size: 11px; margin-top: 15px;'>"
        "Interface Digital Security © 2026</p>",
        unsafe_allow_html=True,
    )

# ---------------------------------------------------------
# CORPO DA APLICAÇÃO
# ---------------------------------------------------------
st.title("📋 Painel da Portaria - Controle de Senhas")

if menu == "📊 Painel de Controle":
    col_titulo, col_btn = st.columns([3, 1])
    with col_titulo:
        st.subheader("Senhas Cadastradas no Sistema")

    with col_btn:
        if os.path.exists(NOME_ARQUIVO):
            with open(NOME_ARQUIVO, "rb") as file:
                st.download_button(
                    label="📥 Baixar Excel",
                    data=file,
                    file_name=f"relatorio_senhas_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_download_painel",
                )

    if not registros:
        st.info("Nenhuma senha cadastrada no momento.")
    else:
        df = pd.DataFrame(registros)
        df = df[
            ["id", "nome", "senha", "status", "data_vencimento", "prazo_dias"]
        ]
        df.columns = [
            "ID",
            "Nome do Responsável",
            "Senha (4 Dígitos)",
            "Status",
            "Data de Vencimento",
            "Prazo (Dias)",
        ]
        df["Data de Vencimento"] = pd.to_datetime(
            df["Data de Vencimento"]
        ).dt.strftime("%d/%m/%Y %H:%M")

        st.dataframe(df, use_container_width=True, hide_index=True)

elif menu == "🔑 Novo Cadastro de Senha":
    st.subheader("Cadastrar Novo Visitante / Prestador")

    with st.form("form_cadastro"):
        col1, col2 = st.columns(2)
        with col1:
            nome_input = st.text_input("Nome:").strip()
        with col2:
            sobrenome_input = st.text_input("Sobrenome:").strip()

        prazo_input = st.number_input(
            "Prazo de Validade em Dias (Máximo 30 dias):",
            min_value=1,
            max_value=30,
            value=30,
            step=1,
        )

        submit = st.form_submit_button("Gerar Senha de Acesso")

    if submit:
        if not nome_input or not sobrenome_input:
            st.warning("⚠️ Preenchimento obrigatório: Digite o Nome e o Sobrenome.")
        else:
            nome_completo = f"{nome_input} {sobrenome_input}".upper().strip()

            nomes_existentes = [
                r["nome"].upper() for r in registros if r["status"] == "Ativa"
            ]

            if nome_completo in nomes_existentes:
                st.error(
                    f"⚠️ O responsável '{nome_completo}' já possui uma senha"
                    " ATIVA cadastrada!"
                )
            else:
                senhas_ativas = {
                    r["senha"] for r in registros if r["status"] == "Ativa"
                }
                while True:
                    nova_senha = f"{random.randint(0, 9999):04d}"
                    if nova_senha not in senhas_ativas:
                        break

                agora = datetime.now()
                vencimento = agora + timedelta(days=int(prazo_input))

                novo_registro = {
                    "nome": nome_completo,
                    "senha": nova_senha,
                    "data_criacao": agora,
                    "prazo_dias": int(prazo_input),
                    "data_vencimento": vencimento,
                    "status": "Ativa",
                }

                registros.append(novo_registro)
                salvar_dados(registros)

                st.success(
                    f"✅ Senha [{nova_senha}] gerada e salva na planilha Excel"
                    f" para '{nome_completo}'."
                )
                st.rerun()

elif menu == "✏️ Editar Cadastro":
    st.subheader("Alterar Dados de Registro Existente")

    if not registros:
        st.info("Nenum cadastro disponível para edição.")
    else:
        opcoes = [
            f"ID {r['id']} - {r['nome']} (Senha: {r['senha']})"
            for r in registros
        ]
        selecionado = st.selectbox("Selecione o registro para editar:", opcoes)

        id_selecionado = int(selecionado.split(" ")[1])
        reg_atual = next(r for r in registros if r["id"] == id_selecionado)

        partes_nome = reg_atual["nome"].split(" ", 1)
        nome_padrao = partes_nome[0]
        sobrenome_padrao = partes_nome[1] if len(partes_nome) > 1 else ""

        with st.form("form_edicao"):
            st.info(f"🔑 **Senha (Não alterável):** `{reg_atual['senha']}`")

            col1, col2 = st.columns(2)
            with col1:
                novo_nome = st.text_input("Nome:", value=nome_padrao).strip()
            with col2:
                novo_sobrenome = st.text_input(
                    "Sobrenome:", value=sobrenome_padrao
                ).strip()

            novo_prazo = st.number_input(
                "Prazo de Validade em Dias (Máximo 30 dias):",
                min_value=1,
                max_value=30,
                value=int(reg_atual["prazo_dias"]),
                step=1,
            )

            submit_edit = st.form_submit_button("Salvar Alterações")

        if submit_edit:
            if not novo_nome or not novo_sobrenome:
                st.warning(
                    "⚠️ Preenchimento obrigatório: Digite o Nome e o"
                    " Sobrenome."
                )
            else:
                novo_nome_completo = (
                    f"{novo_nome} {novo_sobrenome}".upper().strip()
                )
                hoje = datetime.now()

                nova_data_vencimento = reg_atual["data_criacao"] + timedelta(
                    days=int(novo_prazo)
                )
                novo_status = (
                    "Ativa"
                    if int(novo_prazo) > 0 and nova_data_vencimento >= hoje
                    else "Expirada"
                )

                reg_atual["nome"] = novo_nome_completo
                reg_atual["prazo_dias"] = int(novo_prazo)
                reg_atual["data_vencimento"] = nova_data_vencimento
                reg_atual["status"] = novo_status

                salvar_dados(registros)

                st.success(
                    f"✅ Cadastro do ID {id_selecionado} ({novo_nome_completo})"
                    " atualizado com sucesso no Excel!"
                )
                st.rerun()

elif menu == "🗑️ Cancelar / Remover Senha":
    st.subheader("Remover Registro e Liberar Acesso")

    if not registros:
        st.info("Nenhum cadastro disponível no momento.")
    else:
        opcoes = [f"ID {r['id']} - {r['nome']}" for r in registros]
        selecionado = st.selectbox("Selecione o registro para remover:", opcoes)

        if "confirmando_exclusao" not in st.session_state:
            st.session_state.confirmando_exclusao = False

        if not st.session_state.confirmando_exclusao:
            if st.button("Iniciar Exclusão de Registro"):
                st.session_state.confirmando_exclusao = True
                st.rerun()
        else:
            st.warning(
                f"⚠️ **CONFIRMAÇÃO DE SEGURANÇA**: Tem certeza de que deseja apagar permanentemente o registro **{selecionado}**?"
            )

            col_sim, col_nao = st.columns(2)

            with col_sim:
                if st.button("✔️ Sim, Confirmar e Apagar"):
                    id_remover = int(selecionado.split(" ")[1])
                    novos_registros = [
                        r for r in registros if r["id"] != id_remover
                    ]

                    salvar_dados(novos_registros)

                    st.session_state.confirmando_exclusao = False
                    st.success(
                        "✅ Acesso removido e planilha Excel atualizada com"
                        " sucesso!"
                    )
                    st.rerun()

            with col_nao:
                if st.button("❌ Não, Cancelar Operação"):
                    st.session_state.confirmando_exclusao = False
                    st.info("Operação cancelada com segurança.")
                    st.rerun()